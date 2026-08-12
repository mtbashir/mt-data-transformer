"""Reading/writing of the workbook formats used by this project.

Handles .xlsx/.xlsm (openpyxl), .xlsb (pyxlsb) and .csv, plus the header-row
detection needed because some sheets (e.g. Master Data 'All SKU') carry a
title row above the real header.
"""
from __future__ import annotations

import os
import re
import datetime as _dt

import numpy as np
import pandas as pd

XLSB = ".xlsb"
CSV_EXTS = {".csv", ".txt", ".tsv"}
EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xls", ".xlsb"}
ALLOWED_EXTS = EXCEL_EXTS | CSV_EXTS

# Excel's day-zero for serial dates (the 1900 system, including the leap bug).
EXCEL_EPOCH = _dt.datetime(1899, 12, 30)


def ext_of(path: str) -> str:
    return os.path.splitext(path)[1].lower()


def _have_calamine() -> bool:
    try:
        import python_calamine  # noqa: F401
        return True
    except ImportError:
        return False


HAVE_CALAMINE = _have_calamine()


def engine_for(path: str):
    """Fastest available reader for this file type.

    calamine reads a large .xlsx several times faster than openpyxl - on a
    250k-row workbook it is the difference between half a minute and two - so
    prefer it when installed and fall back to openpyxl when it is not.
    """
    e = ext_of(path)
    if e == XLSB:
        return "pyxlsb"
    if e in (".xlsx", ".xlsm", ".xltx") and HAVE_CALAMINE:
        return "calamine"
    return None


def list_sheets(path: str) -> list[str]:
    """Sheet names, without loading any cell data we don't need."""
    if ext_of(path) in CSV_EXTS:
        return ["(csv)"]
    if ext_of(path) == XLSB:
        import pyxlsb

        with pyxlsb.open_workbook(path) as wb:
            return list(wb.sheets)
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _read_raw(path: str, sheet: str | None, nrows: int | None = None) -> pd.DataFrame:
    """Read with no header at all, so we can find the header row ourselves."""
    if ext_of(path) in CSV_EXTS:
        sep = "\t" if ext_of(path) == ".tsv" else None
        return pd.read_csv(
            path, header=None, nrows=nrows, sep=sep, engine="python",
            dtype=object, keep_default_na=True,
        )
    return pd.read_excel(
        path, sheet_name=sheet, header=None, nrows=nrows,
        engine=engine_for(path), dtype=object,
    )


_HEADER_CACHE: dict[tuple, int] = {}


def detect_header_row(path: str, sheet: str | None, scan: int = 12) -> int:
    """Best-guess 0-based index of the header row.

    Scores each candidate row on how header-like it is: mostly non-empty,
    mostly text, no duplicates, and followed by a row that looks like data
    (i.e. contains numbers/dates where the header had text).

    Cached: `nrows` does not stop the underlying readers from parsing the whole
    sheet, so on a large workbook this costs almost as much as a full read.
    """
    hkey = _cache_key(path, sheet, "__header__")
    if hkey is not None and hkey in _HEADER_CACHE:
        return _HEADER_CACHE[hkey]

    raw = _read_raw(path, sheet, nrows=scan + 3)
    if raw.empty:
        return 0

    best, best_score = 0, float("-inf")
    for i in range(min(scan, len(raw))):
        row = raw.iloc[i]
        vals = [v for v in row.tolist() if v is not None and str(v).strip() != ""]
        if not vals:
            continue
        n = len(vals)
        texts = [v for v in vals if isinstance(v, str) and str(v).strip() != ""]
        # A header is mostly text, densely filled, and has unique labels.
        score = 0.0
        score += 2.0 * (len(texts) / n)
        score += 1.5 * (n / max(1, row.notna().sum() or n))
        score += 1.0 * (len(vals) / max(1, raw.shape[1]))
        dupes = n - len({str(v).strip().lower() for v in vals})
        score -= 0.8 * dupes
        # Prefer a row whose successor looks like data rather than more labels.
        if i + 1 < len(raw):
            nxt = [v for v in raw.iloc[i + 1].tolist() if v is not None and str(v).strip() != ""]
            if nxt:
                numeric_next = sum(
                    1 for v in nxt
                    if isinstance(v, (int, float, _dt.datetime, _dt.date))
                    and not isinstance(v, bool)
                )
                score += 1.2 * (numeric_next / len(nxt))
        score -= 0.15 * i  # earlier rows win ties
        if score > best_score:
            best, best_score = i, score

    if hkey is not None:
        if len(_HEADER_CACHE) > 64:
            _HEADER_CACHE.clear()
        _HEADER_CACHE[hkey] = best
    return best


def _dedupe_columns(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out


def _clean_columns(cols) -> list[str]:
    cleaned = []
    for i, c in enumerate(cols):
        if c is None or (isinstance(c, float) and np.isnan(c)) or str(c).strip() == "":
            cleaned.append(f"Column {i + 1}")
        else:
            name = re.sub(r"\s+", " ", str(c).replace("\n", " ")).strip()
            cleaned.append(name or f"Column {i + 1}")
    return _dedupe_columns(cleaned)


# Parsed sheets, keyed by file identity + how it was read. A large workbook
# costs tens of seconds to parse; re-reading it for every reload of the same
# unchanged file is pure waiting. Invalidated by mtime and size.
_CACHE: dict[tuple, pd.DataFrame] = {}
_CACHE_MAX = 8


def _cache_key(path: str, sheet, header_row) -> tuple | None:
    try:
        st = os.stat(path)
    except OSError:
        return None
    return (os.path.realpath(path), st.st_mtime_ns, st.st_size, str(sheet), header_row)


def read_table(path: str, sheet: str | None = None, header_row: int | None = None) -> pd.DataFrame:
    """Load a sheet into a DataFrame with cleaned column names.

    Drops fully-empty rows/columns and any unnamed all-empty leading columns
    (the Master Data 'All SKU' sheet starts with one).
    """
    if sheet in (None, "", "(csv)") and ext_of(path) in CSV_EXTS:
        sheet = None
    if header_row is None:
        header_row = detect_header_row(path, sheet)

    # Returned without copying - copying a 250k-row frame costs seconds. Under
    # pandas' copy-on-write the callers that mutate (build_output, standardise)
    # already take their own copy first, so the cached frame stays clean.
    key = _cache_key(path, sheet, header_row)
    if key is not None and key in _CACHE:
        return _CACHE[key]

    if ext_of(path) in CSV_EXTS:
        sep = "\t" if ext_of(path) == ".tsv" else None
        df = pd.read_csv(path, header=header_row, sep=sep, engine="python")
    else:
        df = pd.read_excel(path, sheet_name=sheet, header=header_row, engine=engine_for(path))

    df.columns = _clean_columns(df.columns)
    df = df.dropna(axis=0, how="all")
    # Drop columns that are entirely empty AND were never really named.
    # pandas labels blank header cells 'Unnamed: N'; we label them 'Column N'.
    unnamed = re.compile(r"(Column \d+|Unnamed:\s*\d+(\.\d+)?)")
    drop = [c for c in df.columns if df[c].isna().all() and unnamed.fullmatch(str(c))]
    if drop:
        df = df.drop(columns=drop)
    df = df.reset_index(drop=True)

    # Dates often arrive as bare Excel serial numbers - always in .xlsb, which
    # has no date type, but also in .xlsx whenever the cells were never given a
    # date format. Convert any column whose name says "date" and whose values
    # sit in a plausible serial range, whatever the file type.
    for c in df.columns:
        if re.search(r"date", str(c), re.I) and pd.api.types.is_numeric_dtype(df[c]):
            s = pd.to_numeric(df[c], errors="coerce")
            valid = s.dropna()
            if len(valid) and valid.between(20000, 80000).mean() > 0.9:
                df[c] = excel_serial_to_datetime(s)

    if key is not None:
        if len(_CACHE) >= _CACHE_MAX:
            _CACHE.pop(next(iter(_CACHE)))
        _CACHE[key] = df
    return df


def excel_serial_to_datetime(s: pd.Series) -> pd.Series:
    return pd.to_datetime(EXCEL_EPOCH) + pd.to_timedelta(pd.to_numeric(s, errors="coerce"), unit="D")


def datetime_to_excel_serial(s: pd.Series) -> pd.Series:
    dt = pd.to_datetime(s, errors="coerce")
    return (dt - pd.Timestamp(EXCEL_EPOCH)).dt.total_seconds() / 86400.0


def json_safe(value):
    """Convert a single pandas/numpy value into something json.dumps accepts."""
    if value is None:
        return None
    if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        f = float(value)
        return None if (np.isnan(f) or np.isinf(f)) else f
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (pd.Timestamp, _dt.datetime, _dt.date)):
        ts = pd.Timestamp(value)
        if ts.time() == _dt.time(0, 0):
            return ts.strftime("%Y-%m-%d")
        return ts.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M:%S")
    return str(value)


def preview(df: pd.DataFrame, rows: int = 25) -> dict:
    """Column metadata + a small sample, shaped for the UI."""
    head = df.head(rows)
    cols = []
    for c in df.columns:
        s = df[c]
        non_null = s.notna().sum()
        cols.append({
            "name": str(c),
            "dtype": dtype_label(s),
            "non_null": int(non_null),
            "null": int(len(s) - non_null),
            "unique": int(s.nunique(dropna=True)),
            "samples": [json_safe(v) for v in s.dropna().head(3).tolist()],
        })
    return {
        "columns": cols,
        "rows": [[json_safe(v) for v in rec] for rec in head.itertuples(index=False, name=None)],
        "row_count": int(len(df)),
        "col_count": int(df.shape[1]),
    }


def dtype_label(s: pd.Series) -> str:
    if pd.api.types.is_datetime64_any_dtype(s):
        return "date"
    if pd.api.types.is_bool_dtype(s):
        return "bool"
    if pd.api.types.is_numeric_dtype(s):
        return "number"
    return "text"


def _cell_value(v):
    """A value openpyxl can write: native types, blanks as None, dates as dates."""
    if v is None or v is pd.NA:
        return None
    if isinstance(v, float) and np.isnan(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.generic):
        return v.item()
    return v


def _as_date_set(s: pd.Series) -> set:
    """The distinct calendar dates in a column, ignoring any time-of-day."""
    d = pd.to_datetime(s, errors="coerce").dropna()
    return set(d.dt.normalize())


def _shared_date_column(df_cols, target_headers: dict) -> str | None:
    """A column named like a date that exists in both the output and the sheet.

    `target_headers` is keyed by casefolded name. Prefer an exact 'date' match,
    then any name containing 'date'.
    """
    pairs = [(str(c), str(c).strip().casefold()) for c in df_cols]
    for name, key in pairs:
        if key == "date" and key in target_headers:
            return name
    for name, key in pairs:
        if "date" in key and key in target_headers:
            return name
    return None


def append_to_excel(df: pd.DataFrame, path: str, sheet: str | None = None,
                    skip_existing_dates: bool = True) -> dict:
    """Append the rows of `df` to an existing sheet, keeping everything else.

    The target workbook must be writable (.xlsx/.xlsm - openpyxl cannot save
    .xlsb/.xls). Rows are aligned to the sheet's own header by name, so column
    order in `df` does not have to match; any column `df` has that the sheet
    lacks is added as a new trailing header. Other sheets, formulas and
    formatting in the workbook are left untouched. Written to a temp file and
    swapped in, so a failure never leaves a half-written original.

    When `skip_existing_dates` is set (the default) and both the output and the
    sheet share a date column, rows whose date is already present in the sheet
    are left out - so re-running never duplicates a date that is already there.
    """
    from openpyxl import load_workbook

    ext = ext_of(path)
    if ext not in (".xlsx", ".xlsm", ".xltx"):
        raise ValueError(
            f"Can only append to .xlsx/.xlsm files; '{os.path.basename(path)}' "
            f"is {ext}. Save it as .xlsx first, or generate a new file instead.")

    wb = load_workbook(path)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif sheet:
        raise ValueError(f"Sheet '{sheet}' is not in {os.path.basename(path)}. "
                         f"It has: {', '.join(wb.sheetnames)}.")
    else:
        ws = wb[wb.sheetnames[0]]

    header_row = detect_header_row(path, ws.title) + 1  # openpyxl rows are 1-based
    headers = {}  # cleaned name -> column index (1-based)
    max_col = 0
    for cell in ws[header_row]:
        max_col = max(max_col, cell.column)
        if cell.value is not None and str(cell.value).strip():
            headers[str(cell.value).strip().casefold()] = cell.column

    # Skip dates the file already holds, so a re-run never doubles a date.
    skipped_rows = 0
    skipped_dates: list[str] = []
    date_col = None
    if skip_existing_dates:
        date_col = _shared_date_column(df.columns, headers)
        if date_col is not None:
            existing_df = read_table(path, ws.title)
            existing = _as_date_set(existing_df[date_col]) \
                if date_col in existing_df.columns else set()
            if existing:
                incoming = pd.to_datetime(df[date_col], errors="coerce").dt.normalize()
                dup = incoming.isin(existing)
                if dup.any():
                    skipped_dates = sorted(
                        {d.strftime("%Y-%m-%d") for d in incoming[dup].dropna().unique()})
                    skipped_rows = int(dup.sum())
                    df = df[~dup]

    # Map each output column to a target column, creating new ones as needed.
    col_index = {}
    for name in df.columns:
        key = str(name).strip().casefold()
        if key in headers:
            col_index[name] = headers[key]
        else:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=str(name))
            headers[key] = max_col
            col_index[name] = max_col

    start = max(ws.max_row, header_row) + 1
    if len(df):
        for r, (_, row) in enumerate(df.iterrows(), start=start):
            for name in df.columns:
                ws.cell(row=r, column=col_index[name], value=_cell_value(row[name]))
        tmp = path + ".tmp"
        wb.save(tmp)
        wb.close()
        os.replace(tmp, path)
    else:
        # Nothing new to write - do not rewrite the file at all.
        wb.close()
    return {"path": path, "sheet": ws.title, "appended": int(len(df)),
            "first_row": start, "skipped": skipped_rows,
            "skipped_dates": skipped_dates, "date_column": date_col}


def write_excel(sheets: dict[str, pd.DataFrame], out_path: str) -> str:
    """Write one or more DataFrames to a formatted .xlsx workbook."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(out_path, engine="openpyxl", datetime_format="yyyy-mm-dd") as xl:
        for name, df in sheets.items():
            safe = re.sub(r"[\[\]\*:\?/\\]", "_", str(name))[:31] or "Sheet1"
            df.to_excel(xl, sheet_name=safe, index=False)
            ws = xl.sheets[safe]
            header_fill = PatternFill("solid", fgColor="1F3864")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            for i, col in enumerate(df.columns, start=1):
                # An all-blank column leaves .max() as NA under pandas' string
                # dtype, and int(NA) raises - so fall back to the header width.
                lengths = df[col].head(400).astype("object").map(
                    lambda v: 0 if v is None or (isinstance(v, float) and np.isnan(v))
                    or v is pd.NA else len(str(v)))
                longest = int(lengths.max()) if len(lengths) and lengths.notna().any() else 0
                width = max(len(str(col)), longest)
                ws.column_dimensions[get_column_letter(i)].width = min(max(width + 3, 10), 42)
            if len(df):
                ws.auto_filter.ref = ws.dimensions
    return out_path
