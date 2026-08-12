"""Idempotent load of the clean output into the Excel 'database' workbook.

Design goals:

* **Idempotent** - re-running the same day must not double-count. Rows are
  keyed (default: DATE + CITY + SKU STANDARD NAME); any existing rows sharing an
  incoming key are removed before the new rows are appended, so a re-run for a
  date *replaces* that date rather than duplicating it.
* **Safe** - the target workbook is copied into a timestamped backup before it
  is touched, so a bad run is always recoverable.
* **Non-destructive to siblings** - other sheets in the workbook are carried
  through unchanged in content.

This module is deliberately self-contained (pandas + openpyxl only) so it can be
unit-tested without the rest of the engine.
"""
from __future__ import annotations

import datetime as _dt
import os
import re
import shutil

import numpy as np
import pandas as pd


def _key_series(df: pd.DataFrame, cols: list[str]) -> pd.Series:
    """One comparable key string per row, stable across dtypes."""
    if not cols:
        return pd.Series([""] * len(df), index=df.index)
    parts = []
    for c in cols:
        s = df[c] if c in df.columns else pd.Series([None] * len(df), index=df.index)
        if pd.api.types.is_datetime64_any_dtype(s):
            t = pd.to_datetime(s, errors="coerce").dt.strftime("%Y-%m-%d")
        elif pd.api.types.is_numeric_dtype(s):
            t = s.map(lambda v: "" if pd.isna(v) else (
                str(int(v)) if float(v).is_integer() else str(float(v))))
        else:
            t = s.astype(object).map(
                lambda v: "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip())
        parts.append(t.astype(str).str.upper())
    out = parts[0]
    for p in parts[1:]:
        out = out + "" + p
    return out


def resolve_keys(df: pd.DataFrame, key_cols: list[str] | None) -> list[str]:
    """Keep only key columns that exist; fall back to sensible defaults."""
    if key_cols:
        present = [c for c in key_cols if c in df.columns]
        # All-or-nothing: partial keys could delete the wrong rows, so if any
        # configured key is missing we fall back to append-only (empty keys).
        return present if len(present) == len(key_cols) else []
    # Default: the natural grain of this data set, if present.
    for candidate in (["DATE", "CITY", "SKU STANDARD NAME"],
                      ["UNIQUE ID"], ["ID"]):
        present = [c for c in candidate if c in df.columns]
        if len(present) == len(candidate):
            return present
    return []


def backup(path: str, backup_dir: str) -> str | None:
    if not os.path.exists(path):
        return None
    os.makedirs(backup_dir, exist_ok=True)
    stamp = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(path))[0]
    dest = os.path.join(backup_dir, f"{base}_{stamp}.xlsx")
    shutil.copyfile(path, dest)
    return dest


def append_idempotent(db_path: str, sheet: str, clean_df: pd.DataFrame,
                      key_cols: list[str] | None = None,
                      backup_dir: str | None = None) -> dict:
    """Merge `clean_df` into `sheet` of the workbook at `db_path`.

    Returns a summary describing what changed.
    """
    summary = {"sheet": sheet, "rows_incoming": int(len(clean_df)),
               "rows_before": 0, "rows_replaced": 0, "rows_after": 0,
               "backup": None, "keys": []}

    backup_dir = backup_dir or os.path.join(os.path.dirname(db_path) or ".", "backups")
    summary["backup"] = backup(db_path, backup_dir)

    # Read every existing sheet so siblings survive the rewrite.
    existing_sheets: dict[str, pd.DataFrame] = {}
    if os.path.exists(db_path):
        try:
            existing_sheets = pd.read_excel(db_path, sheet_name=None)
        except Exception:
            existing_sheets = {}

    prior = existing_sheets.get(sheet)
    keys = resolve_keys(clean_df if prior is None else prior, key_cols)
    summary["keys"] = keys

    if prior is not None and len(prior):
        summary["rows_before"] = int(len(prior))
        # Align incoming columns to the established sheet order.
        merged_cols = list(prior.columns) + [c for c in clean_df.columns
                                             if c not in prior.columns]
        prior = prior.reindex(columns=merged_cols)
        incoming = clean_df.reindex(columns=merged_cols)
        if keys and all(k in prior.columns for k in keys):
            incoming_keys = set(_key_series(incoming, keys))
            prior_key = _key_series(prior, keys)
            keep = ~prior_key.isin(incoming_keys)
            summary["rows_replaced"] = int((~keep).sum())
            prior = prior.loc[keep]
        result = pd.concat([prior, incoming], ignore_index=True)
    else:
        result = clean_df.copy()

    summary["rows_after"] = int(len(result))
    existing_sheets[sheet] = result
    _write_workbook(existing_sheets, db_path)
    return summary


def _write_workbook(sheets: dict[str, pd.DataFrame], out_path: str) -> None:
    """Write all sheets back with a light, consistent header style."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    tmp = out_path + ".tmp.xlsx"
    with pd.ExcelWriter(tmp, engine="openpyxl", datetime_format="yyyy-mm-dd") as xl:
        for name, df in sheets.items():
            safe = re.sub(r"[\[\]\*:\?/\\]", "_", str(name))[:31] or "Sheet1"
            df = df.copy()
            df.to_excel(xl, sheet_name=safe, index=False)
            ws = xl.sheets[safe]
            fill = PatternFill("solid", fgColor="1F3864")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            for i, col in enumerate(df.columns, start=1):
                sample = df[col].head(400).astype(str)
                width = max(len(str(col)), int(sample.str.len().max()) if len(sample) else 0)
                ws.column_dimensions[get_column_letter(i)].width = min(max(width + 3, 10), 42)
            if len(df):
                ws.auto_filter.ref = ws.dimensions
    os.replace(tmp, out_path)
